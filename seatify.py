# *************************************************************************************************************
# Name: Seatify
# Description: Pulls data from spotify to determine most played artist on popular play lists
# Author: Eric Armstrong
# *************************************************************************************************************

from collections import defaultdict
import spotipy
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.colors import WHITE
from openpyxl.worksheet.table import Table, TableStyleInfo
from spotipy.oauth2 import SpotifyClientCredentials


# *************************************************************************************************************
# Function: create_wb
# Description: Creates excel file that displays data of artists appearing in popular Spotify playlists
# Input: data:list, genre:string
# Output: Excel File
# *************************************************************************************************************

def create_wb(data, genre):
    print(f'Creating {genre} spreadsheet...\n')

    wb = Workbook()
    destination = f'{genre}.xlsx'

    sea_green = styles.colors.Color(rgb='69BE28')
    sea_blue = styles.colors.Color(rgb='002244')
    sea_gray = styles.colors.Color(rgb='A5ACAF')

    green_fill = styles.fills.PatternFill(patternType='solid', fgColor=sea_green)
    blue_fill = styles.fills.PatternFill(patternType='solid', fgColor=sea_blue)
    gray_fill = styles.fills.PatternFill(patternType='solid', fgColor=sea_gray)
    font = Font(color=WHITE)

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color=WHITE),
        right=Side(border_style=BORDER_THIN, color=WHITE),
        top=Side(border_style=BORDER_THIN, color=WHITE),
        bottom=Side(border_style=BORDER_THIN, color=WHITE)
    )

    ws1 = wb.active
    ws1.title = genre
    ws1.cell(column=1, row=1, value='Artist').fill = gray_fill
    ws1.cell(column=2, row=1, value='Playlist Entries').fill = gray_fill
    ws1.cell(column=1, row=1).font = font
    ws1.cell(column=2, row=1).font = font

    alt = 1
    for row in range(2, 102):
        if alt < 0:
            ws1.cell(column=1, row=row, value=data[row - 2][0]).fill = green_fill
            ws1.cell(column=1, row=row).font = font
            ws1.cell(column=1, row=row).border = thin_border
            ws1.cell(column=2, row=row, value=data[row - 2][1]).fill = green_fill
            ws1.cell(column=2, row=row).font = font
            ws1.cell(column=2, row=row).border = thin_border
            alt *= -1
        else:
            ws1.cell(column=1, row=row, value=data[row - 2][0]).fill = blue_fill
            ws1.cell(column=1, row=row).font = font
            ws1.cell(column=1, row=row).border = thin_border
            ws1.cell(column=2, row=row, value=data[row - 2][1]).fill = blue_fill
            ws1.cell(column=2, row=row).font = font
            ws1.cell(column=2, row=row).border = thin_border
            alt *= -1

    tab = Table(displayName="Table1", ref="A1:B101")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20
    tab.tableStyleInfo = style
    ws1.add_table(tab)
    wb.save(filename=destination)


# *************************************************************************************************************
# Function: update_dict
# Description: Adds artists to dictionary, and/or updates the amount of times the artist has appeared in
#              playlist category
# Input: data:dictionary(json data), song_dict:dictionary
# Output: None
# *************************************************************************************************************

def update_dict(data, song_dict):
    for i, p in enumerate(data['tracks']['items']):
        if p['track'] is not None:
            if song_dict.get(p['track']['artists'][0]['name']) is not None:
                song_dict[p['track']['artists'][0]['name']] += 1
            else:
                song_dict[p['track']['artists'][0]['name']] = 1


# *************************************************************************************************************
# Function: playlist_tracker():
# Description: Creates Excel files for ranking of popular artists on spotify
# Input: None
# Output: Excel File
# *************************************************************************************************************

def playlist_tracker():
    # initialize credentials, and categories
    client_credentials_manager = SpotifyClientCredentials()
    sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
    categories = ['party', 'hiphop', 'pop']

    # retrieve data from spotify and place sorted data in workbook
    for i in range(len(categories)):
        print(f'Retrieving {categories[i]} playlist data....\n')
        song_dict = defaultdict()

        # response retrieves playlist data from spotify based on category of playlist
        response = sp.category_playlists(limit=50, category_id=f'{categories[i]}', country='US')

        while response:
            playlists = response['playlists']

            for item in (playlists['items']):
                play_id = item['id']
                uri = 'spotify:spotify:spotifycharts:playlist:{}'.format(play_id)
                username = uri.split(':')[2]  # user is spotify...we are using spotify created playlists
                playlist_id = uri.split(':')[4]
                results = sp.user_playlist(username, playlist_id)
                update_dict(results, song_dict)

            if playlists['next']:
                response = sp.next(playlists)
            else:
                response = None

            # sort results
            workbook_list = sorted(song_dict.items(), key=lambda k_v: k_v[1], reverse=True)

            create_wb(workbook_list, categories[i])


playlist_tracker()
